#!/usr/bin/env python3
"""
Workout Dashboard v2 Parser
Reads MacroFactor XLSX exports + Health Connect SQLite DB
Outputs public/data.json

Usage: uv run python parse.py
       HC_DB_PATH=./path/to.db uv run python parse.py
"""

import json
import os
import re
import sqlite3
from datetime import datetime, timezone, date as date_cls, timedelta
from pathlib import Path

import openpyxl

MF_DIR = Path("drive_export/workout")
HC_DB_PATH = Path(os.environ.get("HC_DB_PATH", "health_connect_export.db"))
OUT_PATH = Path("public/data.json")
CONFIG_PATH = Path("workout-config.json")

# 22 muscle groups in display order (matches MF column names without unit suffix)
MUSCLE_GROUPS = [
    "Chest", "Quads", "Upper Back", "Glutes", "Lats", "Hamstrings",
    "Biceps", "Triceps", "Front Delts", "Side Delts", "Lower Back",
    "Abs", "Calves", "Upper Traps", "Rear Delts", "Forearms",
    "Obliques", "Abductors", "Adductors", "Neck", "Tibialis", "Serratus",
]

# Superset suffix pattern: " ∈ SS1", " ∈ C2", etc. — strip " ∈ " + anything after
SS_PATTERN = re.compile(r"\s*∈.*$")

DEFAULT_CONFIG = {
    "athlete": {
        "age": 31,
        "height_cm": 181,
        "waist_cm": 101,
        "start_weight_kg": 120,
        "target_bf_pct": 15,
    },
    # MPS landmarks: calibrated for MF fractional-attribution set counting
    # (compound movements attribute fractional sets to multiple muscles, so weekly
    #  group totals are higher than per-muscle counts in the literature)
    "volume_landmarks": {
        "Push":  {"mev": 40, "mav": 70,  "mrv": 100},
        "Pull":  {"mev": 40, "mav": 70,  "mrv": 100},
        "Upper": {"mev": 80, "mav": 140, "mrv": 200},
        "Lower": {"mev": 25, "mav": 45,  "mrv": 70},
    },
}

# ── MPS (Muscle Performance Set) scale ───────────────────────────────────────
# A quality-weighted set count based on proximity to failure and set type.
# Drop Set = first heavy leg (failure); Drop = continuation leg after weight drop.

MPS_BY_RIR = {0: 1.0, 1: 0.9, 2: 0.75, 3: 0.5, 4: 0.3}

# Muscle groups by movement pattern (for Push/Pull/Upper/Lower indexes)
PUSH_MUSCLES  = {"Chest", "Triceps", "Front Delts", "Side Delts"}
PULL_MUSCLES  = {"Upper Back", "Lats", "Biceps", "Rear Delts"}
UPPER_MUSCLES = PUSH_MUSCLES | PULL_MUSCLES
LOWER_MUSCLES = {"Quads", "Hamstrings", "Glutes", "Calves", "Adductors", "Abductors", "Lower Back"}


# ── Helpers ───────────────────────────────────────────────────────────────────

def to_date_str(val) -> str | None:
    """Convert str, datetime, or date → YYYY-MM-DD string."""
    if val is None:
        return None
    if isinstance(val, str):
        return val[:10] if len(val) >= 10 else None
    if hasattr(val, "strftime"):
        return val.strftime("%Y-%m-%d")
    return None


def clean_exercise(name: str) -> str:
    """Strip superset suffix (e.g. ' ∈ SS1') from exercise name."""
    return SS_PATTERN.sub("", name).strip() if name else ""


def safe_float(val) -> float | None:
    try:
        return float(val) if val is not None else None
    except (ValueError, TypeError):
        return None


def safe_int(val) -> int | None:
    try:
        return int(val) if val is not None else None
    except (ValueError, TypeError):
        return None


# ── MacroFactor XLSX parsing ──────────────────────────────────────────────────

def find_mf_files() -> list[Path]:
    files = sorted(MF_DIR.glob("MacroFactor-*.xlsx"))
    print(f"  Found {len(files)} file(s): {[f.name for f in files]}")
    return files


def parse_quick_export(ws) -> dict:
    """Parse Quick Export sheet → dict keyed by date string."""
    headers = list(next(ws.iter_rows(max_row=1, values_only=True)))
    col = {h: i for i, h in enumerate(headers) if h}

    by_date = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        date = to_date_str(row[0])
        if not date:
            continue
        by_date[date] = {
            "date": date,
            "tdee":              safe_float(row[col["Expenditure"]]) if "Expenditure" in col else None,
            "trend_weight_kg":   safe_float(row[col["Trend Weight (kg)"]]) if "Trend Weight (kg)" in col else None,
            "weight_kg":         safe_float(row[col["Weight (kg)"]]) if "Weight (kg)" in col else None,
            "kcal":              safe_float(row[col["Calories (kcal)"]]) if "Calories (kcal)" in col else None,
            "protein_g":         safe_float(row[col["Protein (g)"]]) if "Protein (g)" in col else None,
            "fat_g":             safe_float(row[col["Fat (g)"]]) if "Fat (g)" in col else None,
            "carbs_g":           safe_float(row[col["Carbs (g)"]]) if "Carbs (g)" in col else None,
            "target_kcal":       safe_float(row[col["Target Calories (kcal)"]]) if "Target Calories (kcal)" in col else None,
            "target_protein_g":  safe_float(row[col["Target Protein (g)"]]) if "Target Protein (g)" in col else None,
            "target_fat_g":      safe_float(row[col["Target Fat (g)"]]) if "Target Fat (g)" in col else None,
            "target_carbs_g":    safe_float(row[col["Target Carbs (g)"]]) if "Target Carbs (g)" in col else None,
            "steps":             safe_int(row[col["Steps"]]) if "Steps" in col else None,
        }
    return by_date


def parse_muscle_sheet(ws) -> dict:
    """
    Parse either Muscle Groups - Sets or Muscle Groups - Volume sheet.
    Column headers like 'Chest (sets)' or 'Chest (kg)' — strip suffix to get muscle name.
    Returns dict keyed by date string.
    """
    headers = list(next(ws.iter_rows(max_row=1, values_only=True)))
    col = {}
    for i, h in enumerate(headers):
        if h and h != "Date":
            muscle = h.replace(" (sets)", "").replace(" (kg)", "").strip()
            col[muscle] = i

    by_date = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        date = to_date_str(row[0])
        if not date:
            continue
        entry = {"date": date}
        for muscle in MUSCLE_GROUPS:
            val = row[col[muscle]] if muscle in col else None
            entry[muscle] = safe_float(val)
        by_date[date] = entry
    return by_date


def parse_workout_log(ws) -> dict:
    """
    Parse Workout Log sheet.
    Groups sets by date → {date, workout_name, duration_sec, sets: [...]}
    """
    headers = list(next(ws.iter_rows(max_row=1, values_only=True)))
    col = {h: i for i, h in enumerate(headers) if h}

    by_date = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        date = to_date_str(row[0])
        if not date:
            continue

        workout_name = row[col.get("Workout", 2)] or ""
        duration_sec = safe_int(row[col.get("Workout Duration", 1)])
        exercise = clean_exercise(row[col.get("Exercise", 3)] or "")
        set_type = row[col.get("Set Type", 5)] or ""
        weight_kg = safe_float(row[col.get("Weight (kg)", 6)])
        reps = safe_int(row[col.get("Reps", 7)])
        rir_raw = row[col.get("RIR", 8)]
        # RIR may be empty string, None, or int
        rir = safe_int(rir_raw) if rir_raw not in (None, "", "None") else None

        set_entry = {
            "exercise": exercise,
            "set_type": set_type,
            "weight_kg": weight_kg,
            "reps": reps,
            "rir": rir,
        }

        if date not in by_date:
            by_date[date] = {
                "date": date,
                "workout_name": workout_name,
                "duration_sec": duration_sec,
                "sets": [],
            }
        by_date[date]["sets"].append(set_entry)

    return by_date


def load_all_mf_files(files: list[Path]) -> dict:
    """Load all MF XLSX files sorted oldest→newest; later files win on duplicate dates."""
    merged = {
        "daily": {},
        "muscle_sets": {},
        "muscle_volume": {},
        "workouts": {},
    }

    for path in files:
        print(f"\n  [{path.name}]")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

        if "Quick Export" in wb.sheetnames:
            d = parse_quick_export(wb["Quick Export"])
            merged["daily"].update(d)
            print(f"    Quick Export       : {len(d)} days")

        if "Muscle Groups - Sets" in wb.sheetnames:
            d = parse_muscle_sheet(wb["Muscle Groups - Sets"])
            merged["muscle_sets"].update(d)
            print(f"    Muscle Groups Sets : {len(d)} days")

        if "Muscle Groups - Volume" in wb.sheetnames:
            d = parse_muscle_sheet(wb["Muscle Groups - Volume"])
            merged["muscle_volume"].update(d)
            print(f"    Muscle Groups Vol  : {len(d)} days")

        if "Workout Log" in wb.sheetnames:
            d = parse_workout_log(wb["Workout Log"])
            merged["workouts"].update(d)
            print(f"    Workout Log        : {len(d)} workout days")

        wb.close()

    return merged


# ── Health Connect parsing ────────────────────────────────────────────────────

def to_hc_date(time_ms: int, zone_s: int) -> str:
    local_ts = (time_ms + (zone_s or 0) * 1000) / 1000
    return datetime.fromtimestamp(local_ts, tz=timezone.utc).strftime("%Y-%m-%d")


def parse_hc_weight(conn: sqlite3.Connection) -> list:
    rows = conn.execute(
        "SELECT weight, time, zone_offset FROM weight_record_table ORDER BY time ASC"
    ).fetchall()
    by_date = {}
    for weight_g, time_ms, zone_s in rows:
        by_date[to_hc_date(time_ms, zone_s or 0)] = round(weight_g / 1000, 2)
    result = [{"date": d, "kg": kg} for d, kg in sorted(by_date.items())]
    print(f"  HC weight : {len(result)} entries")
    return result


def parse_hc_body_fat(conn: sqlite3.Connection) -> list:
    rows = conn.execute(
        "SELECT percentage, time, zone_offset FROM body_fat_record_table ORDER BY time ASC"
    ).fetchall()
    by_date = {}
    for pct, time_ms, zone_s in rows:
        by_date[to_hc_date(time_ms, zone_s or 0)] = round(float(pct), 2)
    result = [{"date": d, "pct": p} for d, p in sorted(by_date.items())]
    print(f"  HC body fat : {len(result)} entries")
    return result


def parse_hc_cardio(conn: sqlite3.Connection) -> list:
    rows = conn.execute("""
        SELECT title, start_time, end_time, start_zone_offset, exercise_type
        FROM exercise_session_record_table
        ORDER BY start_time DESC
    """).fetchall()

    result = []
    for title, start_ms, end_ms, zone_s, ex_type in rows:
        t = (title or "").lower()
        if ex_type == 8 or "vo2" in t or "bike" in t or "cycling" in t:
            session_type = "cardio_vo2"
        elif "padel" in t:
            session_type = "padel"
        else:
            continue
        duration_min = round((end_ms - start_ms) / 1000 / 60)
        result.append({
            "date": to_hc_date(start_ms, zone_s or 0),
            "type": session_type,
            "title": title or "",
            "duration_min": duration_min,
        })

    print(f"  HC cardio : {len(result)} sessions")
    return result


def load_hc_data() -> dict:
    if not HC_DB_PATH.exists():
        print(f"  HC DB not found ({HC_DB_PATH}) — skipping cardio/weight fallback")
        return {"hc_weight": [], "hc_body_fat": [], "cardio": []}

    print(f"  Reading {HC_DB_PATH}...")
    conn = sqlite3.connect(HC_DB_PATH)
    try:
        result = {
            "hc_weight": parse_hc_weight(conn),
            "hc_body_fat": parse_hc_body_fat(conn),
            "cardio": parse_hc_cardio(conn),
        }
    finally:
        conn.close()
    return result


# ── Config ────────────────────────────────────────────────────────────────────

def ensure_config() -> dict:
    if not CONFIG_PATH.exists():
        with open(CONFIG_PATH, "w") as f:
            json.dump(DEFAULT_CONFIG, f, indent=2)
        print(f"  Created {CONFIG_PATH}")
    else:
        print(f"  Loaded  {CONFIG_PATH}")
    with open(CONFIG_PATH) as f:
        return json.load(f)


# ── Merge & output ────────────────────────────────────────────────────────────

def compute_body_comp(weight_series: list, config: dict) -> list:
    """
    Compute BF% estimates (YMCA + Deurenberg), lean mass, FFMI
    for each weight entry that has a trend_kg value.
    Uses fixed waist measurement from config (athlete["waist_cm"]).
    """
    athlete = config["athlete"]
    age = athlete["age"]
    height_m = athlete["height_cm"] / 100.0
    waist_in = athlete["waist_cm"] / 2.54
    bf_manual = athlete.get("bf_pct_manual")  # optional visual/caliper override

    result = []
    for entry in weight_series:
        trend_kg = entry.get("trend_kg")
        if trend_kg is None:
            continue

        weight_lb = trend_kg * 2.20462

        # YMCA formula (male): bf% = ((-98.42 + 4.15*waist_in - 0.082*weight_lb) / weight_lb) * 100
        ymca_bf = ((-98.42 + 4.15 * waist_in - 0.082 * weight_lb) / weight_lb) * 100

        # Deurenberg formula: bf% = (1.20*BMI) + (0.23*age) - 10.8 - 5.4
        bmi = trend_kg / (height_m ** 2)
        deur_bf = (1.20 * bmi) + (0.23 * age) - 10.8 - 5.4

        avg_bf = (ymca_bf + deur_bf) / 2.0

        # If manual BF% set in config, use it as primary (natural waist ≠ navel measurement
        # that formulas require; visual estimate beats formula with wrong input)
        primary_bf = float(bf_manual) if bf_manual is not None else ymca_bf

        lean_kg = trend_kg * (1.0 - primary_bf / 100.0)
        ffmi = lean_kg / (height_m ** 2)

        result.append({
            "date": entry["date"],
            "trend_kg": round(trend_kg, 2),
            "ymca_bf_pct": round(ymca_bf, 1),
            "deurenberg_bf_pct": round(deur_bf, 1),
            "estimated_bf_pct": round(primary_bf, 1),
            "avg_bf_pct": round(avg_bf, 1),
            "lean_kg": round(lean_kg, 1),
            "ffmi": round(ffmi, 2),
        })

    return result


def compute_summary(body_comp: list, mf_daily_list: list = None) -> dict:
    """Pre-compute latest values for the Coach Brief card."""
    if not body_comp:
        return {}

    latest = body_comp[-1]

    # 7-day weight delta
    weight_delta_7d = None
    lean_trend = "→"
    if len(body_comp) >= 7:
        week_ago = body_comp[-7]
        weight_delta_7d = round(latest["trend_kg"] - week_ago["trend_kg"], 2)
        lean_delta = latest["lean_kg"] - week_ago["lean_kg"]
        lean_trend = "↑" if lean_delta > 0.3 else ("↓" if lean_delta < -0.3 else "→")

    # avg_deficit_7d: average (tdee - kcal) over last 7 days where both exist
    avg_deficit_7d = None
    if mf_daily_list:
        today_str = date_cls.today().isoformat()
        cutoff_str = (date_cls.today() - timedelta(days=7)).isoformat()
        deficits = []
        for day in mf_daily_list:
            d = day.get("date", "")
            if d < cutoff_str or d > today_str:
                continue
            tdee = day.get("tdee")
            kcal = day.get("kcal")
            if tdee is not None and kcal is not None:
                deficits.append(tdee - kcal)
        if deficits:
            avg_deficit_7d = round(sum(deficits) / len(deficits), 0)

    return {
        "latest_date": latest["date"],
        "trend_kg": latest["trend_kg"],
        "weight_delta_7d": weight_delta_7d,
        "estimated_bf_pct": latest["estimated_bf_pct"],
        "lean_kg": latest["lean_kg"],
        "lean_trend": lean_trend,
        "ffmi": latest["ffmi"],
        "avg_deficit_7d": avg_deficit_7d,
    }


def compute_cut(body_comp: list, mf_daily_list: list, config: dict) -> dict:
    """Compute cut progress metrics: target weight, rate of loss, projected completion."""
    athlete = config["athlete"]
    start_weight = athlete["start_weight_kg"]
    target_bf = athlete["target_bf_pct"] / 100.0

    if not body_comp:
        return {}

    latest = body_comp[-1]
    lean_kg = latest["lean_kg"]
    current_trend = latest["trend_kg"]
    target_weight = round(lean_kg / (1 - target_bf), 1)
    kg_lost = round(start_weight - current_trend, 1)
    kg_remaining = round(current_trend - target_weight, 1)

    # Rate of loss: linear regression over last 30 body_comp entries with trend_kg
    recent = [e for e in body_comp if e.get("trend_kg") is not None][-30:]
    rate_kg_per_week = None
    projected_date = None
    if len(recent) >= 7:
        n = len(recent)
        xs = list(range(n))
        ys = [e["trend_kg"] for e in recent]
        mean_x = sum(xs) / n
        mean_y = sum(ys) / n
        num = sum((x - mean_x) * (y - mean_y) for x, y in zip(xs, ys))
        den = sum((x - mean_x) ** 2 for x in xs)
        slope_per_day = num / den if den else 0  # kg per day
        rate_kg_per_week = round(slope_per_day * 7, 2)

        if slope_per_day < -0.05 / 7:  # losing at least 0.05 kg/week
            days_to_target = kg_remaining / abs(slope_per_day)
            projected = date_cls.today() + timedelta(days=days_to_target)
            projected_date = projected.strftime("%Y-%m-%d")

    # Count high-deficit days in last 7
    today_str = date_cls.today().isoformat()
    cutoff_str = (date_cls.today() - timedelta(days=7)).isoformat()
    high_deficit_days = 0
    for day in mf_daily_list:
        d = day.get("date", "")
        if d < cutoff_str or d > today_str:
            continue
        tdee = day.get("tdee")
        kcal = day.get("kcal")
        if tdee is not None and kcal is not None:
            deficit = tdee - kcal
            if deficit > 800:
                high_deficit_days += 1

    return {
        "start_weight_kg": start_weight,
        "current_trend_kg": current_trend,
        "target_weight_kg": target_weight,
        "kg_lost": kg_lost,
        "kg_remaining": kg_remaining,
        "rate_kg_per_week": rate_kg_per_week,
        "projected_completion_date": projected_date,
        "high_deficit_days_last7": high_deficit_days,
    }


def set_mps(s: dict) -> float:
    """Return MPS weight for a single set based on set_type and RIR."""
    st = s.get("set_type", "")
    rir = s.get("rir")
    if st == "Drop":          return 0.2   # continuation leg after weight drop
    if st in ("Failure Set", "Drop Set"): return 1.0
    if rir is None:           return 0.75  # unknown effort — conservative
    return MPS_BY_RIR.get(rir, 0.1)        # RIR 5+ → warm-up territory


def compute_mps_by_date(workouts: dict) -> dict:
    """Per-day average MPS multiplier derived from Workout Log set quality."""
    result = {}
    for date, workout in workouts.items():
        sets = workout.get("sets", [])
        if not sets:
            result[date] = 1.0
            continue
        total_mps  = sum(set_mps(s) for s in sets)
        total_sets = len(sets)
        result[date] = round(total_mps / total_sets, 3) if total_sets else 1.0
    return result


def get_iso_week(date_str: str) -> tuple[str, str]:
    """Return (ISO-week label e.g. '2026-W07', monday date string)."""
    d = date_cls.fromisoformat(date_str)
    year, week, _ = d.isocalendar()
    monday = d - timedelta(days=d.weekday())
    return f"{year}-W{week:02d}", monday.strftime("%Y-%m-%d")


def _zone(value: float, lm: dict) -> str:
    mev = lm.get("mev", 0)
    mav = lm.get("mav", (mev + lm.get("mrv", 9999)) / 2)
    mrv = lm.get("mrv", 9999)
    if value < mev:  return "below"
    if value < mav:  return "ok"
    if value <= mrv: return "high"
    return "over"


def compute_push_pull_weekly(muscle_sets: dict, mps_by_date: dict, config: dict) -> list:
    """Aggregate MPS-weighted sets per ISO week into Push/Pull/Upper/Lower totals."""
    lm = config.get("volume_landmarks", {})
    by_week: dict[str, dict] = {}

    for date in sorted(muscle_sets):
        entry  = muscle_sets[date]
        mult   = mps_by_date.get(date, 1.0)
        wk, ws = get_iso_week(date)
        if wk not in by_week:
            by_week[wk] = {"week": wk, "week_start": ws,
                           "push_mps": 0.0, "pull_mps": 0.0,
                           "upper_mps": 0.0, "lower_mps": 0.0,
                           "push_raw": 0.0, "pull_raw": 0.0,
                           "upper_raw": 0.0, "lower_raw": 0.0,
                           "training_days": 0}
        by_week[wk]["training_days"] += 1

        for muscle in MUSCLE_GROUPS:
            raw = entry.get(muscle) or 0.0
            mps = raw * mult
            if muscle in PUSH_MUSCLES:
                by_week[wk]["push_raw"] += raw;  by_week[wk]["push_mps"]  += mps
            if muscle in PULL_MUSCLES:
                by_week[wk]["pull_raw"] += raw;  by_week[wk]["pull_mps"]  += mps
            if muscle in UPPER_MUSCLES:
                by_week[wk]["upper_raw"] += raw; by_week[wk]["upper_mps"] += mps
            if muscle in LOWER_MUSCLES:
                by_week[wk]["lower_raw"] += raw; by_week[wk]["lower_mps"] += mps

    muscle_lm = config.get("muscle_landmarks", {})
    # Per-muscle weekly MPS accumulator alongside group totals
    for wk in by_week:
        by_week[wk]["muscle_mps"]  = {m: 0.0 for m in MUSCLE_GROUPS}
        by_week[wk]["muscle_raw"]  = {m: 0.0 for m in MUSCLE_GROUPS}

    for date in sorted(muscle_sets):
        entry  = muscle_sets[date]
        mult   = mps_by_date.get(date, 1.0)
        wk, _  = get_iso_week(date)
        for muscle in MUSCLE_GROUPS:
            raw = entry.get(muscle) or 0.0
            by_week[wk]["muscle_mps"][muscle]  += raw * mult
            by_week[wk]["muscle_raw"][muscle]  += raw

    result = []
    for wk in sorted(by_week):
        w = by_week[wk]
        push  = round(w["push_mps"],  1)
        pull  = round(w["pull_mps"],  1)
        upper = round(w["upper_mps"], 1)
        lower = round(w["lower_mps"], 1)

        muscles_out = {}
        for muscle in MUSCLE_GROUPS:
            mps = round(w["muscle_mps"][muscle], 1)
            muscles_out[muscle] = {
                "mps":  mps,
                "raw":  round(w["muscle_raw"][muscle], 1),
                "zone": _zone(mps, muscle_lm.get(muscle, {})),
            }

        result.append({
            "week":            wk,
            "week_start":      w["week_start"],
            "training_days":   w["training_days"],
            "push_mps":        push,
            "pull_mps":        pull,
            "upper_mps":       upper,
            "lower_mps":       lower,
            "push_raw":        round(w["push_raw"],  1),
            "pull_raw":        round(w["pull_raw"],  1),
            "upper_raw":       round(w["upper_raw"], 1),
            "lower_raw":       round(w["lower_raw"], 1),
            "push_pull_ratio": round(push / pull, 2) if pull > 0 else None,
            "push_zone":       _zone(push,  lm.get("Push",  {})),
            "pull_zone":       _zone(pull,  lm.get("Pull",  {})),
            "upper_zone":      _zone(upper, lm.get("Upper", {})),
            "lower_zone":      _zone(lower, lm.get("Lower", {})),
            "muscles":         muscles_out,
        })
    return result


def build_weight_series(mf_daily: dict, hc_weight: list) -> list:
    """
    Unified weight series: MF trend weight preferred; HC raw weight as fallback.
    """
    hc_by_date = {e["date"]: e["kg"] for e in hc_weight}
    all_dates = sorted(set(mf_daily.keys()) | set(hc_by_date.keys()))
    result = []
    for date in all_dates:
        mf = mf_daily.get(date)
        trend = mf.get("trend_weight_kg") if mf else None
        raw_mf = mf.get("weight_kg") if mf else None
        hc_kg = hc_by_date.get(date)
        result.append({
            "date": date,
            "trend_kg": trend,
            "raw_kg": raw_mf if raw_mf is not None else hc_kg,
            "source": "mf" if raw_mf is not None else ("hc" if hc_kg else "none"),
        })
    return result


def build_output(mf: dict, hc: dict, config: dict) -> dict:
    mf_daily = mf["daily"]
    weight_series = build_weight_series(mf_daily, hc["hc_weight"])
    body_comp = compute_body_comp(weight_series, config)

    # Enrich each mf_daily entry with deficit field
    mf_daily_list = []
    for d in sorted(mf_daily):
        entry = dict(mf_daily[d])
        tdee = entry.get("tdee")
        kcal = entry.get("kcal")
        entry["deficit"] = round(tdee - kcal, 0) if tdee is not None and kcal is not None else None
        mf_daily_list.append(entry)

    cut = compute_cut(body_comp, mf_daily_list, config)

    workouts_dict = mf["workouts"]
    mps_by_date   = compute_mps_by_date(workouts_dict)
    push_pull_weekly = compute_push_pull_weekly(mf["muscle_sets"], mps_by_date, config)

    return {
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "config": config,
        "summary": compute_summary(body_comp, mf_daily_list),
        "cut": cut,
        "mf_daily": mf_daily_list,
        "muscle_sets": [mf["muscle_sets"][d] for d in sorted(mf["muscle_sets"])],
        "muscle_volume": [mf["muscle_volume"][d] for d in sorted(mf["muscle_volume"])],
        "workouts": [workouts_dict[d] for d in sorted(workouts_dict)],
        "push_pull_weekly": push_pull_weekly,
        "weight": weight_series,
        "body_comp": body_comp,
        "hc_body_fat": hc["hc_body_fat"],
        "cardio": hc["cardio"],
    }


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    print("=== Workout Dashboard v2 Parser ===\n")

    print("Config...")
    config = ensure_config()

    print("\nMacroFactor XLSX files...")
    files = find_mf_files()
    if not files:
        print("  ✗ No MacroFactor files found in drive_export/workout/")
        return
    mf = load_all_mf_files(files)

    print("\nHealth Connect data...")
    hc = load_hc_data()

    print("\nAssembling output...")
    data = build_output(mf, hc, config)

    with open(OUT_PATH, "w") as f:
        json.dump(data, f, separators=(",", ":"))

    size_kb = OUT_PATH.stat().st_size / 1024
    s = data.get("summary", {})
    print(f"\n✓ {OUT_PATH}  ({size_kb:.1f} KB)")
    print(f"  mf_daily         : {len(data['mf_daily'])} days")
    print(f"  muscle_sets      : {len(data['muscle_sets'])} days")
    print(f"  muscle_volume    : {len(data['muscle_volume'])} days")
    print(f"  workouts         : {len(data['workouts'])} days")
    print(f"  push_pull_weekly : {len(data['push_pull_weekly'])} weeks")
    print(f"  body_comp      : {len(data['body_comp'])} entries")
    print(f"  weight         : {len(data['weight'])} entries")
    print(f"  hc_body_fat    : {len(data['hc_body_fat'])} entries")
    print(f"  cardio         : {len(data['cardio'])} sessions")
    if s:
        delta = f"{s['weight_delta_7d']:+.2f}" if s.get("weight_delta_7d") is not None else "n/a"
        print(f"\n  Latest ({s['latest_date']}):")
        print(f"    trend weight : {s['trend_kg']} kg  (7d Δ {delta} kg)")
        print(f"    lean mass    : {s['lean_kg']} kg  {s['lean_trend']}")
        print(f"    BF% (est)    : {s['estimated_bf_pct']}%")
        print(f"    FFMI         : {s['ffmi']}")


if __name__ == "__main__":
    main()
